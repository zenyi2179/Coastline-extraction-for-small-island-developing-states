import os
import pandas as pd

# 定义文件夹路径和目标Excel文件路径
folder_path = r"E:\_OrderingProject\F_IslandsBoundaryChange\f_Python\NCX_JINGHE\analyze\分形维数"
output_excel_path = r"E:\_OrderingProject\F_IslandsBoundaryChange\SIDS_SV\SIDS_SV_v1_compile\Statistics\fractal_dimension\fractal_dimension.xlsx"

# 获取文件夹下所有csv文件
csv_files = [f for f in os.listdir(folder_path) if f.endswith('.csv')]

# 创建一个ExcelWriter对象，用于将多个DataFrame写入到一个Excel文件中
with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
    # 初始化一个空的DataFrame用于存储汇总数据
    summary_df = pd.DataFrame()

    for csv_file in csv_files:
        # 获取不带扩展名的文件名，作为工作表名称
        sheet_name = os.path.splitext(csv_file)[0]
        sheet_name = sheet_name.split('FD_')[-1]

        # 构造完整的文件路径
        file_path = os.path.join(folder_path, csv_file)

        # 读取csv文件内容到DataFrame
        df = pd.read_csv(file_path)

        # 将DataFrame写入到Excel文件中的一个工作表中
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # 准备汇总数据
        # 假设每个csv文件的结构如下：
        # grid_size_m,occupied_grids,fractal_dimension
        # 12000,1587,1.123637691
        # 4800,4541,1.123637691
        # 2400,10040,1.123637691
        # 900,30158,1.123637691
        # 300,99935,1.123637691
        # 提取fractal_dimension值
        fractal_dimension = df['fractal_dimension'].iloc[0]

        # 创建一个临时DataFrame用于存储当前文件的汇总数据
        temp_df = pd.DataFrame({
            'dataset': [sheet_name],
            'fractal_dimension': [fractal_dimension],
            '12000': [df[df['grid_size_m'] == 12000]['occupied_grids'].values[0]],
            '4800': [df[df['grid_size_m'] == 4800]['occupied_grids'].values[0]],
            '2400': [df[df['grid_size_m'] == 2400]['occupied_grids'].values[0]],
            '900': [df[df['grid_size_m'] == 900]['occupied_grids'].values[0]],
            '300': [df[df['grid_size_m'] == 300]['occupied_grids'].values[0]]
        })

        # 将临时DataFrame添加到汇总DataFrame中
        summary_df = pd.concat([summary_df, temp_df], ignore_index=True)

    # 将汇总DataFrame写入到Excel文件中的一个名为'All'的工作表中
    summary_df.to_excel(writer, sheet_name='All', index=False)

# 重新加载Excel文件并调整工作表顺序
from openpyxl import load_workbook

# 加载Excel文件
wb = load_workbook(output_excel_path)

# 获取所有工作表名称
sheet_names = wb.sheetnames

# 将'All'工作表移动到第一个位置
if 'All' in sheet_names:
    sheet_names.remove('All')
    sheet_names.insert(0, 'All')

# 重新排序工作表
wb._sheets = [wb[sheet_name] for sheet_name in sheet_names]

# 保存调整后的工作表顺序
wb.save(output_excel_path)

print(
    f"所有csv文件已分别保存到 {output_excel_path} 的不同工作表中，并且汇总数据已保存到'All'工作表中，且'All'工作表已提前到第一个显示")